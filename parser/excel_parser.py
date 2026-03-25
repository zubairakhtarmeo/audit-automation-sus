from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter


SERIAL_HEADER_KEYWORDS = {
    "sl",
    "sl#",
    "sl #",
    "s.l",
    "s.l.",
    "sr",
    "sr#",
    "sno",
    "s/no",
    "serial",
    "serial no",
    "serial number",
    "no",
    "no.",
    "#",
}

QUESTION_HEADER_KEYWORDS = {
    "question",
    "questions",
    "requirement",
    "requirements",
    "audit point",
    "audit points",
    "check point",
    "checkpoint",
    "compliance point",
}

ANSWER_HEADER_KEYWORDS = {
    "present status",
    "present\nstatus",
    "current status",
    "status",
    "response",
    "answer",
    "remarks",
    "comment",
    "comments",
    "observation",
    "observations",
    "finding",
    "findings",
    "evidence",
}

SCORE_HEADER_KEYWORDS = {
    "score",
    "marks",
    "mark",
    "rating",
    "points",
    "marks-10",
    "marks - 10",
}

PHOTO_HEADER_KEYWORDS = {
    "picture",
    "photo",
    "evidence",
    "image",
    "snapshot",
    "supporting document",
    "supporting documents",
}

IMPROVEMENT_HEADER_KEYWORDS = {
    "improvement",
    "improvement plan",
    "action plan",
    "cap",
    "corrective action",
    "corrective action plan",
    "improvement plan (cap)",
}

REFERENCE_SHEET_KEYWORDS = {
    "legend",
    "instruction",
    "instructions",
    "guideline",
    "guidelines",
    "notes",
    "note",
    "abbreviation",
    "abbreviations",
    "definitions",
    "definition",
    "scoring",
    "scoring guide",
    "how to",
}


@dataclass(frozen=True)
class SheetColumnMap:
    header_row: Optional[int]
    sl_col: Optional[int]
    question_col: Optional[int]
    answer_col: Optional[int]
    score_col: Optional[int]
    photo_col: Optional[int]
    improvement_col: Optional[int]
    criteria_cols: Dict[str, int]


def parse_excel(file_path: str) -> List[Dict[str, Any]]:
    """Parse a brand audit workbook (.xlsx) into a normalized question list.

    This parser is intentionally heuristic-driven so it can handle many different
    brand formats without hardcoding per-brand logic.

    Returns a list of dicts like:
        {
          "sheet": "Latest update",
          "row_index": 12,
          "sl": "1",
          "question": "Is there an IE department?",
          "criteria": {"10": "...", "5": "..."},
          "answer_col": "D",
          "score_col": "E",
          "photo_col": "F",
          "improvement_col": "G",
        }

    Notes:
    - row_index is 1-based Excel row index.
    - Column fields are Excel letters ("A", "B", ...). They may be None if not detected.
    """

    wb = load_workbook(filename=file_path, data_only=True)
    results: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        merged_anchor_map = _build_merged_anchor_map(ws)

        # Quick sheet classification; skip obvious reference/legend sheets.
        if _looks_like_reference_sheet(ws, merged_anchor_map):
            continue

        colmap = _detect_columns_openpyxl(ws, merged_anchor_map)
        sheet_questions = _extract_questions_openpyxl(ws, merged_anchor_map, colmap)

        # pandas fallback for flat sheets if we found nothing meaningful
        if not sheet_questions:
            try:
                sheet_questions = _extract_questions_pandas(file_path, ws.title)
            except Exception:
                sheet_questions = []

        results.extend(sheet_questions)

    return results


def parse_audit_file(file_path: str) -> List[Dict[str, Any]]:
    """Compatibility wrapper for the UI.

    The Streamlit module imports `parse_audit_file` as the public entrypoint.
    """

    return parse_excel(file_path)


# -----------------------
# openpyxl helpers
# -----------------------

def _build_merged_anchor_map(ws) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """Map every cell inside merged ranges -> (top_left_row, top_left_col)."""

    anchor_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        anchor = (min_row, min_col)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                anchor_map[(r, c)] = anchor
    return anchor_map


def _cell_anchor(merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]], row: int, col: int) -> Tuple[int, int]:
    return merged_anchor_map.get((row, col), (row, col))


def _get_cell_value(ws, merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]], row: int, col: int) -> Any:
    ar, ac = _cell_anchor(merged_anchor_map, row, col)
    return ws.cell(row=ar, column=ac).value


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd is not None and pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, str):
        text = value
    else:
        text = str(value)

    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _norm_header(value: Any) -> str:
    text = _norm_text(value).lower()
    text = text.replace(":", "").strip()
    return text


def _is_empty_row(ws, merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]], row: int, max_col: int) -> bool:
    for c in range(1, max_col + 1):
        if _norm_text(_get_cell_value(ws, merged_anchor_map, row, c)):
            return False
    return True


def _looks_like_reference_sheet(ws, merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]]) -> bool:
    name = (ws.title or "").strip().lower()
    if any(k in name for k in ("legend", "instruction", "guideline", "notes", "definitions")):
        return True

    max_row = min(ws.max_row or 1, 30)
    max_col = min(ws.max_column or 1, 25)

    text_blob = " ".join(
        _norm_header(_get_cell_value(ws, merged_anchor_map, r, c))
        for r in range(1, max_row + 1)
        for c in range(1, max_col + 1)
        if _norm_header(_get_cell_value(ws, merged_anchor_map, r, c))
    )

    if not text_blob:
        return True

    hits = sum(1 for kw in REFERENCE_SHEET_KEYWORDS if kw in text_blob)

    # If it looks like a reference page and does not show a question structure, skip it.
    if hits >= 6 and ("question" not in text_blob):
        return True

    return False


def _detect_header_row(ws, merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]]) -> Optional[int]:
    max_row = min(ws.max_row or 1, 60)
    max_col = min(ws.max_column or 1, 40)

    best_row: Optional[int] = None
    best_score = 0

    for r in range(1, max_row + 1):
        row_texts = [_norm_header(_get_cell_value(ws, merged_anchor_map, r, c)) for c in range(1, max_col + 1)]
        row_texts = [t for t in row_texts if t]
        if not row_texts:
            continue

        row_blob = " | ".join(row_texts)

        score = 0
        if any(any(k == t or k in t for k in SERIAL_HEADER_KEYWORDS) for t in row_texts):
            score += 3
        if any(any(k in t for k in QUESTION_HEADER_KEYWORDS) for t in row_texts):
            score += 4
        if any(any(k in t for k in ANSWER_HEADER_KEYWORDS) for t in row_texts):
            score += 2
        if any(any(k in t for k in SCORE_HEADER_KEYWORDS) for t in row_texts):
            score += 2
        if any(any(k in t for k in PHOTO_HEADER_KEYWORDS) for t in row_texts):
            score += 1
        if any(any(k in t for k in IMPROVEMENT_HEADER_KEYWORDS) for t in row_texts):
            score += 1
        if any(re.fullmatch(r"(10|5|3|0)", t) for t in row_texts):
            score += 2
        if "week 2" in row_blob or "wk 2" in row_blob:
            score += 1

        if score > best_score:
            best_score = score
            best_row = r

    if best_score >= 4:
        return best_row
    return None


def _pick_best_candidate(headers_by_col: Dict[int, str], candidates: Sequence[int]) -> Optional[int]:
    if not candidates:
        return None

    def penalty(col_idx: int) -> Tuple[int, int]:
        h = headers_by_col.get(col_idx, "")
        p = 0
        if "week 2" in h or "wk 2" in h or "w2" in h:
            p += 3
        if "week" in h and "week 1" not in h and "wk 1" not in h:
            p += 1
        # Secondary tie-breaker: earlier columns are usually the primary set
        return (p, col_idx)

    return min(candidates, key=penalty)


def _detect_columns_openpyxl(ws, merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]]) -> SheetColumnMap:
    header_row = _detect_header_row(ws, merged_anchor_map)
    max_col = min(ws.max_column or 1, 60)

    headers_by_col: Dict[int, str] = {}
    if header_row is not None:
        for c in range(1, max_col + 1):
            headers_by_col[c] = _norm_header(_get_cell_value(ws, merged_anchor_map, header_row, c))

    sl_candidates: List[int] = []
    question_candidates: List[int] = []
    answer_candidates: List[int] = []
    score_candidates: List[int] = []
    photo_candidates: List[int] = []
    improvement_candidates: List[int] = []
    criteria_cols: Dict[str, int] = {}

    if header_row is not None:
        for c, h in headers_by_col.items():
            if not h:
                continue

            if any(h == kw or kw in h for kw in SERIAL_HEADER_KEYWORDS):
                sl_candidates.append(c)

            if any(kw in h for kw in QUESTION_HEADER_KEYWORDS):
                question_candidates.append(c)

            if any(kw in h for kw in ANSWER_HEADER_KEYWORDS):
                answer_candidates.append(c)

            if any(kw in h for kw in SCORE_HEADER_KEYWORDS):
                score_candidates.append(c)

            if any(kw in h for kw in PHOTO_HEADER_KEYWORDS):
                photo_candidates.append(c)

            if any(kw in h for kw in IMPROVEMENT_HEADER_KEYWORDS):
                improvement_candidates.append(c)

            # Criteria columns sometimes are just 10/5/3/0 headers
            m = re.fullmatch(r"(10|5|3|0)", h)
            if m:
                criteria_cols[m.group(1)] = c
            else:
                # Variants like "10 marks" or "10 points"
                m2 = re.match(r"^(10|5|3|0)\b", h)
                if m2 and ("mark" in h or "point" in h or "score" in h or "criteria" in h):
                    criteria_cols[m2.group(1)] = c

    sl_col = _pick_best_candidate(headers_by_col, sl_candidates)
    question_col = _pick_best_candidate(headers_by_col, question_candidates)
    answer_col = _pick_best_candidate(headers_by_col, answer_candidates)
    score_col = _pick_best_candidate(headers_by_col, score_candidates)
    photo_col = _pick_best_candidate(headers_by_col, photo_candidates)
    improvement_col = _pick_best_candidate(headers_by_col, improvement_candidates)

    # Data-driven fallback when header-based detection fails.
    if header_row is None or (sl_col is None and question_col is None):
        inferred = _infer_columns_by_data(ws, merged_anchor_map, start_row=(header_row or 1))
        sl_col = sl_col or inferred.sl_col
        question_col = question_col or inferred.question_col
        answer_col = answer_col or inferred.answer_col
        score_col = score_col or inferred.score_col
        photo_col = photo_col or inferred.photo_col
        improvement_col = improvement_col or inferred.improvement_col
        criteria_cols = criteria_cols or inferred.criteria_cols

    return SheetColumnMap(
        header_row=header_row,
        sl_col=sl_col,
        question_col=question_col,
        answer_col=answer_col,
        score_col=score_col,
        photo_col=photo_col,
        improvement_col=improvement_col,
        criteria_cols=criteria_cols,
    )


def _infer_columns_by_data(ws, merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]], start_row: int) -> SheetColumnMap:
    max_row = min(ws.max_row or 1, start_row + 120)
    max_col = min(ws.max_column or 1, 40)

    numeric_counts = {c: 0 for c in range(1, max_col + 1)}
    seq_counts = {c: 0 for c in range(1, max_col + 1)}
    text_len_sum = {c: 0 for c in range(1, max_col + 1)}
    text_counts = {c: 0 for c in range(1, max_col + 1)}

    last_num_by_col: Dict[int, Optional[int]] = {c: None for c in range(1, max_col + 1)}

    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            v = _get_cell_value(ws, merged_anchor_map, r, c)
            txt = _norm_text(v)
            if not txt:
                continue

            num = _try_parse_int(txt)
            if num is not None:
                numeric_counts[c] += 1
                prev = last_num_by_col.get(c)
                if prev is not None and num == prev + 1:
                    seq_counts[c] += 1
                last_num_by_col[c] = num
            else:
                # consider it text
                if len(txt) >= 6:
                    text_len_sum[c] += len(txt)
                    text_counts[c] += 1

    # Serial column tends to have lots of small integers and a sequence.
    sl_col = max(numeric_counts.keys(), key=lambda c: (seq_counts[c], numeric_counts[c])) if numeric_counts else None
    if sl_col is not None and (numeric_counts[sl_col] < 4 and seq_counts[sl_col] < 2):
        sl_col = None

    # Question column tends to have long text.
    def avg_text_len(c: int) -> float:
        if text_counts[c] == 0:
            return 0.0
        return text_len_sum[c] / text_counts[c]

    question_col = max(text_len_sum.keys(), key=avg_text_len) if text_len_sum else None
    if question_col is not None and avg_text_len(question_col) < 12:
        question_col = None

    return SheetColumnMap(
        header_row=None,
        sl_col=sl_col,
        question_col=question_col,
        answer_col=None,
        score_col=None,
        photo_col=None,
        improvement_col=None,
        criteria_cols={},
    )


def _try_parse_int(text: str) -> Optional[int]:
    t = text.strip()
    if not t:
        return None

    # Common patterns: "1", "1.", "01", "1)"
    m = re.match(r"^(\d{1,4})(?:\s*[\).#-])?\s*$", t)
    if not m:
        return None

    try:
        return int(m.group(1))
    except Exception:
        return None


def _is_probable_question_text(text: str) -> bool:
    if not text:
        return False

    t = text.strip()

    # Exclude obvious headers/legends
    low = t.lower()
    if any(k in low for k in ("legend", "instruction", "guideline", "scoring", "definitions")):
        return False

    # Exclude column headers repeated mid-sheet
    if any(k in low for k in ("present status", "marks", "score", "picture", "improvement plan", "response")) and len(t) < 40:
        return False

    alpha = sum(1 for ch in t if ch.isalpha())
    if alpha < 4:
        return False

    if len(t) < 8:
        return False

    return True


def _infer_question_text_from_row(
    ws,
    merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]],
    row: int,
    max_col: int,
    preferred_col: Optional[int],
    exclude_cols: Sequence[int],
) -> str:
    # Prefer explicit question column if detected
    if preferred_col is not None:
        txt = _norm_text(_get_cell_value(ws, merged_anchor_map, row, preferred_col))
        if _is_probable_question_text(txt):
            return txt

    # Otherwise: pick the longest meaningful text cell in the row.
    best = ""
    for c in range(1, max_col + 1):
        if c in exclude_cols:
            continue
        txt = _norm_text(_get_cell_value(ws, merged_anchor_map, row, c))
        if not _is_probable_question_text(txt):
            continue
        if len(txt) > len(best):
            best = txt

    return best


def _extract_questions_openpyxl(ws, merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]], colmap: SheetColumnMap) -> List[Dict[str, Any]]:
    max_col = min(ws.max_column or 1, 80)
    max_row = ws.max_row or 1

    start_row = (colmap.header_row + 1) if colmap.header_row is not None else 1

    # Stop after too many consecutive empty rows, to avoid scanning thousands of blank rows.
    consecutive_empty = 0

    seen_anchors: set[Tuple[int, int]] = set()
    rows_out: List[Dict[str, Any]] = []

    for r in range(start_row, max_row + 1):
        if _is_empty_row(ws, merged_anchor_map, r, min(max_col, 30)):
            consecutive_empty += 1
            if consecutive_empty >= 50 and rows_out:
                break
            continue
        consecutive_empty = 0

        # Serial number
        sl_text = ""
        if colmap.sl_col is not None:
            sl_text = _norm_text(_get_cell_value(ws, merged_anchor_map, r, colmap.sl_col))
        sl_num = _try_parse_int(sl_text) if sl_text else None

        # Infer a serial if not mapped
        if sl_num is None:
            for c in range(1, min(max_col, 10) + 1):
                sl_candidate = _norm_text(_get_cell_value(ws, merged_anchor_map, r, c))
                parsed = _try_parse_int(sl_candidate)
                if parsed is not None:
                    sl_num = parsed
                    sl_text = str(parsed)
                    break

        # Question text
        exclude_cols = [
            colmap.sl_col or -1,
            colmap.answer_col or -1,
            colmap.score_col or -1,
            colmap.photo_col or -1,
            colmap.improvement_col or -1,
            *list(colmap.criteria_cols.values()),
        ]
        question = _infer_question_text_from_row(
            ws,
            merged_anchor_map,
            row=r,
            max_col=min(max_col, 50),
            preferred_col=colmap.question_col,
            exclude_cols=[c for c in exclude_cols if c != -1],
        )

        # Skip non-question rows.
        if sl_num is None and not question:
            continue
        if sl_num is None:
            # Some sheets omit serials; allow questions if question text is strong.
            if not _is_probable_question_text(question):
                continue
        else:
            # When a serial exists, still require a plausible question.
            if not _is_probable_question_text(question):
                continue

        # Deduplicate merged question blocks
        if colmap.question_col is not None:
            ar, ac = _cell_anchor(merged_anchor_map, r, colmap.question_col)
            if (ar, ac) in seen_anchors:
                continue
            seen_anchors.add((ar, ac))
            row_index = ar
        else:
            row_index = r

        criteria: Dict[str, str] = {}
        for k, c in colmap.criteria_cols.items():
            criteria[k] = _norm_text(_get_cell_value(ws, merged_anchor_map, row_index, c))

        rows_out.append(
            {
                "sheet": ws.title,
                "row_index": int(row_index),
                "sl": sl_text.strip() if sl_text else (str(sl_num) if sl_num is not None else ""),
                "question": question,
                "criteria": criteria,
                "answer_col": get_column_letter(colmap.answer_col) if colmap.answer_col else None,
                "score_col": get_column_letter(colmap.score_col) if colmap.score_col else None,
                "photo_col": get_column_letter(colmap.photo_col) if colmap.photo_col else None,
                "improvement_col": get_column_letter(colmap.improvement_col) if colmap.improvement_col else None,
            }
        )

    return rows_out


# -----------------------
# pandas fallback
# -----------------------

def _extract_questions_pandas(file_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    if pd is None:
        raise ImportError("pandas is not installed; pandas fallback is unavailable")

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=object)
    if df.empty:
        return []

    # Drop fully empty rows/cols
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if df.empty:
        return []

    # Find a header row using similar heuristics
    max_scan = min(len(df), 60)
    best_row = None
    best_score = 0

    for i in range(max_scan):
        row_vals = [_norm_header(v) for v in df.iloc[i].tolist()]
        row_vals = [v for v in row_vals if v]
        if not row_vals:
            continue

        blob = " | ".join(row_vals)
        score = 0
        if any(any(k == t or k in t for k in SERIAL_HEADER_KEYWORDS) for t in row_vals):
            score += 3
        if any(any(k in t for k in QUESTION_HEADER_KEYWORDS) for t in row_vals):
            score += 4
        if any(any(k in t for k in ANSWER_HEADER_KEYWORDS) for t in row_vals):
            score += 2
        if any(re.fullmatch(r"(10|5|3|0)", t) for t in row_vals):
            score += 2
        if "week 2" in blob or "wk 2" in blob:
            score += 1

        if score > best_score:
            best_score = score
            best_row = i

    header_idx = best_row if best_score >= 4 else None

    headers_by_col: Dict[int, str] = {}
    if header_idx is not None:
        for c in range(df.shape[1]):
            headers_by_col[c] = _norm_header(df.iat[header_idx, c])

    def find_col(keywords: Iterable[str]) -> List[int]:
        cols: List[int] = []
        for c, h in headers_by_col.items():
            if not h:
                continue
            if any(kw in h or h == kw for kw in keywords):
                cols.append(c)
        return cols

    sl_col = None
    q_col = None
    ans_col = None
    score_col = None
    photo_col = None
    impr_col = None
    criteria_cols: Dict[str, int] = {}

    if header_idx is not None:
        sl_col = _pick_best_candidate({k + 1: v for k, v in headers_by_col.items()}, [c + 1 for c in find_col(SERIAL_HEADER_KEYWORDS)])
        q_col = _pick_best_candidate({k + 1: v for k, v in headers_by_col.items()}, [c + 1 for c in find_col(QUESTION_HEADER_KEYWORDS)])
        ans_col = _pick_best_candidate({k + 1: v for k, v in headers_by_col.items()}, [c + 1 for c in find_col(ANSWER_HEADER_KEYWORDS)])
        score_col = _pick_best_candidate({k + 1: v for k, v in headers_by_col.items()}, [c + 1 for c in find_col(SCORE_HEADER_KEYWORDS)])
        photo_col = _pick_best_candidate({k + 1: v for k, v in headers_by_col.items()}, [c + 1 for c in find_col(PHOTO_HEADER_KEYWORDS)])
        impr_col = _pick_best_candidate({k + 1: v for k, v in headers_by_col.items()}, [c + 1 for c in find_col(IMPROVEMENT_HEADER_KEYWORDS)])

        for c, h in headers_by_col.items():
            m = re.fullmatch(r"(10|5|3|0)", h)
            if m:
                criteria_cols[m.group(1)] = c + 1
            else:
                m2 = re.match(r"^(10|5|3|0)\b", h)
                if m2 and ("mark" in h or "point" in h or "score" in h or "criteria" in h):
                    criteria_cols[m2.group(1)] = c + 1

    start = (header_idx + 1) if header_idx is not None else 0
    out: List[Dict[str, Any]] = []

    for i in range(start, len(df)):
        row_vals = df.iloc[i].tolist()

        sl_text = ""
        sl_num = None
        if sl_col is not None:
            sl_text = _norm_text(row_vals[sl_col - 1])
            sl_num = _try_parse_int(sl_text) if sl_text else None
        if sl_num is None:
            # scan first few columns for serial
            for c in range(min(10, len(row_vals))):
                t = _norm_text(row_vals[c])
                parsed = _try_parse_int(t)
                if parsed is not None:
                    sl_num = parsed
                    sl_text = str(parsed)
                    break

        # question
        question = ""
        if q_col is not None:
            question = _norm_text(row_vals[q_col - 1])
        if not _is_probable_question_text(question):
            # longest meaningful
            best = ""
            for v in row_vals:
                t = _norm_text(v)
                if _is_probable_question_text(t) and len(t) > len(best):
                    best = t
            question = best

        if sl_num is None and not question:
            continue
        if question and not _is_probable_question_text(question):
            continue

        criteria: Dict[str, str] = {}
        for k, c in criteria_cols.items():
            if 1 <= c <= len(row_vals):
                criteria[k] = _norm_text(row_vals[c - 1])

        out.append(
            {
                "sheet": sheet_name,
                "row_index": int(i + 1),
                "sl": sl_text.strip() if sl_text else (str(sl_num) if sl_num is not None else ""),
                "question": question,
                "criteria": criteria,
                "answer_col": get_column_letter(ans_col) if ans_col else None,
                "score_col": get_column_letter(score_col) if score_col else None,
                "photo_col": get_column_letter(photo_col) if photo_col else None,
                "improvement_col": get_column_letter(impr_col) if impr_col else None,
            }
        )

    return out
