from __future__ import annotations

import os
from copy import copy
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string


@dataclass(frozen=True)
class WriteResult:
    output_path: str
    written_cells: int
    skipped_existing: int
    skipped_invalid: int
    skipped_missing_sheet: int


def write_filled_workbook(
    input_path: str,
    answered_questions: List[Dict[str, Any]],
) -> WriteResult:
    """Write AI answers back into the original workbook, preserving formatting.

    `answered_questions` is expected to be a list of dicts like:
      {
        "sheet": "Latest update",
        "row_index": 12,
        "answer_col": "D",
        "score_col": "E",
        "photo_col": "F",
        "improvement_col": "G",
        "present_status": "...",
        "score": "10",
        "improvement_plan": "...",
      }

    Behavior:
    - Never overwrites non-empty cells.
    - If a target cell is inside a merged range, writes to the merged range's top-left cell.
    - Applies Arial 10 + wrap_text for cells we write.
    - Writes "To be provided" into photo/evidence cell if empty.
    - Saves as Filled_[original_filename].xlsx next to input.

    Returns summary stats and output path.
    """

    wb = load_workbook(filename=input_path, data_only=False)

    written_cells = 0
    skipped_existing = 0
    skipped_invalid = 0
    skipped_missing_sheet = 0

    for item in answered_questions:
        sheet_name = str(item.get("sheet") or "").strip()
        if not sheet_name or sheet_name not in wb.sheetnames:
            skipped_missing_sheet += 1
            continue

        ws = wb[sheet_name]

        row_index = _safe_int(item.get("row_index"))
        if row_index is None or row_index <= 0:
            skipped_invalid += 1
            continue

        # Validate row index: don't write if far outside existing sheet content.
        # openpyxl would extend the sheet, but the requirement says validate.
        if ws.max_row and row_index > ws.max_row:
            skipped_invalid += 1
            continue

        merged_anchor_map = _build_merged_anchor_map(ws)

        # Determine cell coordinates
        answer_col = _safe_col_letter(item.get("answer_col"))
        score_col = _safe_col_letter(item.get("score_col"))
        photo_col = _safe_col_letter(item.get("photo_col"))
        improvement_col = _safe_col_letter(item.get("improvement_col"))

        present_status = _string_or_empty(item.get("present_status"))
        score = _string_or_empty(item.get("score"))
        improvement_plan = _string_or_empty(item.get("improvement_plan"))

        # Present Status
        if answer_col and present_status.strip():
            ok, wrote = _write_if_empty(
                ws,
                merged_anchor_map,
                row_index,
                answer_col,
                present_status,
            )
            if ok:
                if wrote:
                    written_cells += 1
                else:
                    skipped_existing += 1
            else:
                skipped_invalid += 1

        # Score
        if score_col and score.strip():
            ok, wrote = _write_if_empty(
                ws,
                merged_anchor_map,
                row_index,
                score_col,
                score,
            )
            if ok:
                if wrote:
                    written_cells += 1
                else:
                    skipped_existing += 1
            else:
                skipped_invalid += 1

        # Improvement Plan
        if improvement_col and improvement_plan.strip():
            ok, wrote = _write_if_empty(
                ws,
                merged_anchor_map,
                row_index,
                improvement_col,
                improvement_plan,
            )
            if ok:
                if wrote:
                    written_cells += 1
                else:
                    skipped_existing += 1
            else:
                skipped_invalid += 1

        # Photo / evidence column
        if photo_col:
            ok, wrote = _write_if_empty(
                ws,
                merged_anchor_map,
                row_index,
                photo_col,
                "To be provided",
            )
            if ok:
                if wrote:
                    written_cells += 1
                else:
                    skipped_existing += 1
            else:
                skipped_invalid += 1

    output_path = _build_output_path(input_path)
    wb.save(output_path)

    return WriteResult(
        output_path=output_path,
        written_cells=written_cells,
        skipped_existing=skipped_existing,
        skipped_invalid=skipped_invalid,
        skipped_missing_sheet=skipped_missing_sheet,
    )


def write_answers(input_path: str, answered_questions: List[Dict[str, Any]]) -> WriteResult:
    """Compatibility wrapper for the UI."""

    return write_filled_workbook(input_path=input_path, answered_questions=answered_questions)


# -----------------------
# Internals
# -----------------------

def _build_output_path(input_path: str) -> str:
    directory = os.path.dirname(os.path.abspath(input_path))
    base = os.path.basename(input_path)
    name, ext = os.path.splitext(base)
    if not ext:
        ext = ".xlsx"
    return os.path.join(directory, f"Filled_{name}{ext}")


def _safe_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None


def _string_or_empty(value: Any) -> str:
    if value is None:
        return ""
    s = str(value)
    return s


def _safe_col_letter(value: Any) -> Optional[str]:
    if value is None:
        return None
    col = str(value).strip().upper()
    if not col:
        return None
    if not col.isalpha():
        return None
    try:
        # validate by converting to index
        column_index_from_string(col)
    except Exception:
        return None
    return col


def _build_merged_anchor_map(ws) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """Map every cell inside merged ranges -> (top_left_row, top_left_col)."""

    anchor_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        anchor = (min_row, min_col)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                anchor_map[(r, c)] = anchor
    return anchor_map


def _anchor_cell(
    merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]],
    row: int,
    col: int,
) -> Tuple[int, int]:
    return merged_anchor_map.get((row, col), (row, col))


def _cell_is_empty(cell_value: Any) -> bool:
    if cell_value is None:
        return True
    if isinstance(cell_value, str) and cell_value.strip() == "":
        return True
    return False


def _apply_answer_style(cell) -> None:
    # Preserve all other styling aspects; adjust only font name/size and wrap.
    new_font = copy(cell.font) if cell.font else Font()
    new_font.name = "Arial"
    new_font.sz = 10
    cell.font = new_font

    new_alignment = copy(cell.alignment) if cell.alignment else Alignment()
    new_alignment.wrap_text = True
    cell.alignment = new_alignment


def _write_if_empty(
    ws,
    merged_anchor_map: Dict[Tuple[int, int], Tuple[int, int]],
    row_index: int,
    col_letter: str,
    value: str,
) -> Tuple[bool, bool]:
    """Attempt to write value to (row_index, col_letter).

    Returns (ok, wrote):
      - ok=False indicates invalid coordinate.
      - wrote=False indicates cell already had content (not overwritten).
    """

    try:
        col_idx = column_index_from_string(col_letter)
    except Exception:
        return False, False

    ar, ac = _anchor_cell(merged_anchor_map, row_index, col_idx)

    # If row index is outside sheet bounds, reject.
    if ar <= 0 or ac <= 0:
        return False, False

    cell = ws.cell(row=ar, column=ac)

    # Never overwrite existing content
    if not _cell_is_empty(cell.value):
        return True, False

    # Never write empty values
    if value is None or str(value).strip() == "":
        return True, False

    cell.value = value
    _apply_answer_style(cell)
    return True, True
