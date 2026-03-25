from __future__ import annotations

import os
import time
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import load_workbook

from engine.ai_engine import answer_question
from parser.excel_parser import parse_audit_file
from writer.excel_writer import WriteResult, write_answers


APP_TITLE = "Audit Automation Tool"
APP_SUBTITLE = "AI-Powered Compliance Audit Assistant"
MODEL_BADGE = "claude-sonnet-4-6"


def _work_dir() -> Path:
    d = Path(__file__).resolve().parent / "_work"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _save_uploaded_file(uploaded) -> Path:
    work = _work_dir()
    safe_name = Path(uploaded.name).name
    out_path = work / safe_name
    out_path.write_bytes(uploaded.getbuffer())
    return out_path


def _estimate_seconds_remaining(total: int, done: int) -> int:
    remaining = max(total - done, 0)
    return int(remaining * 3)


def _format_seconds(seconds: int) -> str:
    seconds = max(int(seconds), 0)
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    if h:
        return f"{h}h {m}m {s}s"
    if m:
        return f"{m}m {s}s"
    return f"{s}s"


def _sheet_type(parsed_items: List[Dict[str, Any]]) -> str:
    for item in parsed_items:
        criteria = item.get("criteria")
        if isinstance(criteria, dict):
            for k in ("10", "5", "3", "0"):
                v = str(criteria.get(k) or "").strip()
                if v:
                    return "Scored"
    return "Yes/No"


def _build_preview_rows(parsed: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], int]:
    by_sheet: Dict[str, List[Dict[str, Any]]] = {}
    for item in parsed:
        sheet = str(item.get("sheet") or "").strip() or "(Unknown)"
        by_sheet.setdefault(sheet, []).append(item)

    rows: List[Dict[str, Any]] = []
    for sheet, items in sorted(by_sheet.items(), key=lambda x: x[0].lower()):
        rows.append(
            {
                "Sheet Name": sheet,
                "Questions": len(items),
                "Type": _sheet_type(items),
                "Status": "Ready",
            }
        )

    return rows, sum(len(v) for v in by_sheet.values())


def _score_bucket(score: str) -> str:
    s = (score or "").strip()
    return s if s in {"10", "5", "3", "0"} else ""


def _set_api_key_in_env(api_key: str) -> None:
    api_key = (api_key or "").strip()
    if api_key:
        os.environ["ANTHROPIC_API_KEY"] = api_key
    else:
        os.environ.pop("ANTHROPIC_API_KEY", None)


def _render_step_header(num: str, title: str, subtitle: str = "") -> None:
    st.markdown(
        f"""
        <div style="margin-bottom:16px;">
            <div style="font-family:'DM Mono',monospace;
                        font-size:11px; color:#5A5A72;
                        letter-spacing:2px;
                        text-transform:uppercase;
                        margin-bottom:4px;">
                STEP {num}
            </div>
            <div style="font-size:20px; font-weight:500;
                        color:#F1F0FF; margin-bottom:4px;">
                {title}
            </div>
            <div style="font-size:13px; color:#9898B0;">
                {subtitle}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _pill(text: str, kind: str) -> str:
    # kind: ready | scored | yesno
    styles = {
        "ready": "background:#0D2518; border:1px solid #10B981; color:#10B981;",
        "scored": "background:#1C1C26; border:1px solid #6366F1; color:#C7C7FF;",
        "yesno": "background:#2D2010; border:1px solid #F59E0B; color:#F59E0B;",
        "indigo": "background:#1C1C26; border:1px solid #6366F1; color:#F1F0FF;",
    }
    css = styles.get(kind, styles["indigo"])
    return (
        f"<span style=\"display:inline-flex;align-items:center;gap:6px;{css}"
        "border-radius:999px;padding:2px 10px;font-size:12px;"
        "font-family:'DM Sans',sans-serif;\">"
        f"{text}</span>"
    )


def _render_alert(kind: str, title: str, message: str) -> None:
    # kind: error | warning | success | info
    styles = {
        "error": ("#2D1515", "#EF4444", "⚠️"),
        "warning": ("#2D2010", "#F59E0B", "⚠️"),
        "success": ("#0D2518", "#10B981", "✓"),
        "info": ("#16161E", "#6366F1", "ℹ️"),
    }
    bg, border, icon = styles.get(kind, styles["info"])
    st.markdown(
        f"""
        <div style="background:{bg};border:1px solid {border};border-radius:12px;padding:14px 16px;">
            <div style="display:flex;gap:10px;align-items:flex-start;">
                <div style="font-family:'DM Mono',monospace;color:{border};margin-top:1px;">{icon}</div>
                <div>
                    <div style="color:#F1F0FF;font-weight:500;font-size:13px;">{title}</div>
                    <div style="color:#9898B0;font-size:12px;margin-top:2px;white-space:pre-wrap;">{message}</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _render_preview_table(rows: List[Dict[str, Any]]) -> None:
    # Custom HTML table for premium look
    header = """
    <div style="overflow:hidden;border:1px solid #2A2A3A;border-radius:10px;">
      <table style="width:100%;border-collapse:collapse;font-size:13px;">
        <thead>
          <tr style="background:#111118;color:#9898B0;text-align:left;">
            <th style="padding:10px 12px;border-bottom:1px solid #2A2A3A;">Sheet Name</th>
            <th style="padding:10px 12px;border-bottom:1px solid #2A2A3A;">Questions</th>
            <th style="padding:10px 12px;border-bottom:1px solid #2A2A3A;">Type</th>
            <th style="padding:10px 12px;border-bottom:1px solid #2A2A3A;">Status</th>
          </tr>
        </thead>
        <tbody>
    """

    body_parts: List[str] = []
    for i, r in enumerate(rows):
        bg = "#16161E" if i % 2 == 0 else "#1C1C26"
        sheet = str(r.get("Sheet Name") or "")
        questions = int(r.get("Questions") or 0)
        typ = str(r.get("Type") or "")
        status = str(r.get("Status") or "")

        type_badge = _pill("Scored", "scored") if typ == "Scored" else _pill("Yes/No", "yesno")
        status_badge = _pill(status, "ready")

        body_parts.append(
            f"""
            <tr style="background:{bg};border-bottom:1px solid #2A2A3A;">
              <td style="padding:10px 12px;color:#F1F0FF;">{sheet}</td>
              <td style="padding:10px 12px;color:#F1F0FF;font-family:'DM Mono',monospace;">{questions}</td>
              <td style="padding:10px 12px;">{type_badge}</td>
              <td style="padding:10px 12px;">{status_badge}</td>
            </tr>
            """
        )

    footer = """
        </tbody>
      </table>
    </div>
    """

    st.markdown(header + "".join(body_parts) + footer, unsafe_allow_html=True)


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")

    # Required CSS injection (exact block as provided)
    st.markdown(
        """
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500&family=DM+Mono:wght@400;500&display=swap');

/* Hide Streamlit chrome */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
.stDeployButton {display: none;}

/* Main container */
.main .block-container {
    padding: 2rem 2.5rem;
    max-width: 900px;
}

/* Typography */
html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: #111118;
    border-right: 1px solid #2A2A3A;
}
[data-testid="stSidebar"] .block-container {
    padding: 1.5rem 1rem;
}

/* Cards */
.audit-card {
    background: #16161E;
    border: 1px solid #2A2A3A;
    border-radius: 12px;
    padding: 24px;
    margin-bottom: 16px;
}

/* Step indicator */
.step-num {
    font-family: 'DM Mono', monospace;
    font-size: 11px;
    color: #5A5A72;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: 4px;
}

/* Buttons */
.stButton > button {
    background: #6366F1;
    color: white;
    border: none;
    border-radius: 8px;
    height: 48px;
    font-family: 'DM Sans', sans-serif;
    font-weight: 500;
    font-size: 15px;
    width: 100%;
    transition: background 0.2s;
}
.stButton > button:hover {
    background: #818CF8;
    color: white;
    border: none;
}

/* File uploader */
[data-testid="stFileUploader"] {
    background: #16161E;
    border: 1.5px dashed #6366F1;
    border-radius: 12px;
    padding: 20px;
}

/* Progress bar */
.stProgress > div > div {
    background: linear-gradient(90deg, #6366F1, #8B5CF6);
    border-radius: 4px;
}

/* Metrics */
[data-testid="metric-container"] {
    background: #16161E;
    border: 1px solid #2A2A3A;
    border-radius: 10px;
    padding: 16px;
}

/* Input fields */
.stTextInput > div > div > input {
    background: #1C1C26;
    border: 1px solid #2A2A3A;
    border-radius: 8px;
    color: #F1F0FF;
    font-family: 'DM Mono', monospace;
    font-size: 13px;
}

/* Dataframe */
.stDataFrame {
    border: 1px solid #2A2A3A;
    border-radius: 8px;
}

/* Success/Error messages */
.stSuccess {
    background: #0D2518;
    border: 1px solid #10B981;
    border-radius: 8px;
    color: #10B981;
}
.stError {
    background: #2D1515;
    border: 1px solid #EF4444;
    border-radius: 8px;
}
.stWarning {
    background: #2D2010;
    border: 1px solid #F59E0B;
    border-radius: 8px;
}

/* Divider */
hr {
    border-color: #2A2A3A;
}

/* Scrollbar */
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #0A0A0F; }
::-webkit-scrollbar-thumb {
    background: #6366F1;
    border-radius: 3px;
}
</style>
        """,
        unsafe_allow_html=True,
    )

    # Additional polish styles (premium dark background, badges, download button, spinner)
    st.markdown(
        """
<style>
html, body, .stApp {
    background: #0A0A0F;
    color: #F1F0FF;
}

/* Thin top border status bar */
.stApp:before {
    content: '';
    position: fixed;
    top: 0;
    left: 0;
    right: 0;
    height: 2px;
    background: #6366F1;
    z-index: 1000;
}

/* Card hover */
.audit-card:hover {
    border-color: #3A3A52;
    background: #1C1C26;
}

/* Download button as green */
.download-button .stDownloadButton > button {
    background: #10B981 !important;
}
.download-button .stDownloadButton > button:hover {
    background: #34D399 !important;
}

/* Sidebar glass-ish */
[data-testid="stSidebar"] {
    background: rgba(17,17,24,0.92);
    backdrop-filter: blur(10px);
}

/* Spinner */
.spin {
  display:inline-block;
  animation: spin 1s linear infinite;
  font-family: 'DM Mono', monospace;
  color: #6366F1;
}
@keyframes spin { from { transform: rotate(0deg); } to { transform: rotate(360deg);} }

/* Gradient divider */
.gradient-divider {
  height: 1px;
  background: linear-gradient(90deg, rgba(99,102,241,0.0), rgba(99,102,241,1), rgba(139,92,246,1), rgba(99,102,241,0.0));
  margin: 14px 0 18px 0;
}

/* Info bar */
.info-bar {
  background:#111118;
  border:1px solid #2A2A3A;
  border-radius:10px;
  padding:10px 12px;
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap:12px;
}
.info-left { color:#F1F0FF; font-size:13px; }
.info-right { color:#10B981; font-size:12px; font-family:'DM Mono',monospace; }

/* Empty state */
.empty-wrap {
  background:#16161E;
  border:1px dashed #2A2A3A;
  border-radius:12px;
  padding:48px 24px;
  text-align:center;
}
.empty-emoji { font-size:34px; margin-bottom:10px; }
.empty-title { color:#F1F0FF; font-size:16px; font-weight:500; }
.empty-sub { color:#9898B0; font-size:13px; margin-top:4px; }

/* Make download button match other buttons sizing */
.stDownloadButton > button {
    border-radius: 8px;
    height: 48px;
    font-family: 'DM Sans', sans-serif;
    font-weight: 500;
    font-size: 15px;
    width: 100%;
}

/* Reduce default table/df clutter */
div[data-testid="stDataFrame"] > div {
  border-radius: 10px;
}
</style>
        """,
        unsafe_allow_html=True,
    )

    # Sidebar (logo + api key + session stats)
    st.sidebar.markdown(
        """
<div style="display:flex; align-items:center;
            gap:10px; margin-bottom:24px;">
    <div style="width:36px; height:36px;
                background:#6366F1; border-radius:8px;
                display:flex; align-items:center;
                justify-content:center;
                font-weight:500; font-size:13px;
                color:white; font-family:'DM Sans';">
        MG
    </div>
    <div>
        <div style="color:#F1F0FF; font-size:14px;
                    font-weight:500;">
            Audit Automation
        </div>
        <div style="color:#5A5A72; font-size:11px;">
            MG Apparel · v1.0
        </div>
    </div>
</div>
        """,
        unsafe_allow_html=True,
    )
    st.sidebar.markdown("<hr />", unsafe_allow_html=True)

    st.sidebar.markdown(
        "<div style='color:#9898B0;font-size:12px;margin-bottom:8px;'>Anthropic API Key</div>",
        unsafe_allow_html=True,
    )
    api_key = st.sidebar.text_input(
        "Enter Anthropic API Key",
        type="password",
        label_visibility="collapsed",
        value=st.session_state.get("api_key", ""),
    )
    st.session_state["api_key"] = api_key
    _set_api_key_in_env(api_key)

    connected = bool((api_key or "").strip())
    dot_color = "#10B981" if connected else "#EF4444"
    dot_text = "Connected" if connected else "Not connected"
    st.sidebar.markdown(
        f"""
        <div style="display:flex;align-items:center;gap:8px;margin-top:8px;">
            <div style="width:8px;height:8px;border-radius:50%;background:{dot_color};"></div>
            <div style="color:#9898B0;font-size:12px;">{dot_text}</div>
        </div>
        <div style="color:#5A5A72;font-size:11px;margin-top:6px;">Your key is never stored permanently</div>
        """,
        unsafe_allow_html=True,
    )

    st.sidebar.markdown("<hr />", unsafe_allow_html=True)

    files_processed = int(st.session_state.get("files_processed", 0) or 0)
    questions_answered = int(st.session_state.get("questions_answered", 0) or 0)
    if files_processed or questions_answered:
        st.sidebar.markdown(
            "<div style='color:#9898B0;font-size:12px;margin-bottom:10px;'>Session Stats</div>",
            unsafe_allow_html=True,
        )
        st.sidebar.markdown(
            f"""
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;">
              <div style="background:#16161E;border:1px solid #2A2A3A;border-radius:10px;padding:12px;">
                <div style="font-family:'DM Mono',monospace;color:#F1F0FF;font-size:16px;">{files_processed}</div>
                <div style="color:#9898B0;font-size:11px;">Files Processed</div>
              </div>
              <div style="background:#16161E;border:1px solid #2A2A3A;border-radius:10px;padding:12px;">
                <div style="font-family:'DM Mono',monospace;color:#F1F0FF;font-size:16px;">{questions_answered}</div>
                <div style="color:#9898B0;font-size:11px;">Questions Answered</div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.sidebar.markdown("<hr />", unsafe_allow_html=True)

    st.sidebar.markdown(
        "<div style='color:#5A5A72;font-size:11px;'>Audit Automation Tool · v1.0</div>",
        unsafe_allow_html=True,
    )

    # Main header section
    header_left, header_right = st.columns([0.75, 0.25])
    with header_left:
        st.markdown(
            """
            <div style="font-size:34px;font-weight:300;color:#F1F0FF;line-height:1.1;">Audit Automation Tool</div>
            <div style="color:#9898B0;font-size:14px;margin-top:8px;">Upload any brand compliance file. AI fills it in seconds.</div>
            """,
            unsafe_allow_html=True,
        )
    with header_right:
        st.markdown(
            f"""
            <div style="display:flex;justify-content:flex-end;align-items:flex-start;">
              {_pill(MODEL_BADGE, 'indigo')}
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<div class='gradient-divider'></div>", unsafe_allow_html=True)

    # STEP 1 — Upload card
    st.markdown("<div class='audit-card'>", unsafe_allow_html=True)
    _render_step_header(
        "01",
        "Upload Audit File",
        "Supports any brand format — Bestseller, Jack & Jones, Kiabi and more",
    )

    st.markdown(
        "<div style='color:#9898B0;font-size:13px;margin-bottom:10px;'>📂 Drop your .xlsx file here</div>",
        unsafe_allow_html=True,
    )
    uploaded = st.file_uploader("Upload Brand Audit File", type=["xlsx"], label_visibility="collapsed")

    if uploaded is None:
        st.markdown(
            """
            <div class="empty-wrap">
              <div class="empty-emoji">📋</div>
              <div class="empty-title">No file uploaded yet</div>
              <div class="empty-sub">Start by uploading a brand audit file</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)
        return

    if not uploaded.name.lower().endswith(".xlsx"):
        _render_alert("error", "Wrong file format", "Please upload an .xlsx file.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    try:
        input_path = _save_uploaded_file(uploaded)
        st.session_state["input_path"] = str(input_path)
    except Exception as exc:
        _render_alert("error", "Upload failed", f"Failed to save uploaded file.\n{exc}")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    # Parse once per upload
    parsed: List[Dict[str, Any]] = []
    parse_error: Optional[Exception] = None

    if st.session_state.get("parsed_questions") and st.session_state.get("parsed_for") == str(input_path):
        parsed = st.session_state["parsed_questions"]
    else:
        with st.spinner("Analyzing workbook structure..."):
            try:
                parsed = parse_audit_file(str(input_path))
                st.session_state["parsed_questions"] = parsed
                st.session_state["parsed_for"] = str(input_path)
            except Exception as exc:
                parse_error = exc

    if parse_error is not None:
        _render_alert("error", "Parsing failed", f"Could not read questions from this Excel file.\n{parse_error}")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    # Basic upload info
    sheet_count = 0
    try:
        # Read-only load just to count sheets; does not affect formatting.
        tmp_wb = load_workbook(filename=str(input_path), read_only=True, data_only=True)
        sheet_count = len(tmp_wb.sheetnames)
        tmp_wb.close()
    except Exception:
        # Fallback to parsed sheet list if workbook cannot be opened read-only
        sheet_count = len({str(q.get("sheet") or "").strip() for q in parsed if q.get("sheet")})

        st.markdown(
                f"""
                <div class="info-bar" style="margin-top:12px;">
                    <div class="info-left">
                        <span style="font-family:'DM Mono',monospace;color:#C7C7FF;">{uploaded.name}</span>
                        <span style="color:#5A5A72;">&nbsp;|&nbsp;</span>
                        <span style="font-family:'DM Mono',monospace;">{sheet_count} sheets</span>
                        <span style="color:#5A5A72;">&nbsp;|&nbsp;</span>
                        <span style="color:#9898B0;">Ready to process</span>
                    </div>
                    <div class="info-right">✓ READY</div>
                </div>
                """,
                unsafe_allow_html=True,
        )

        st.markdown("</div>", unsafe_allow_html=True)

        # STEP 2 — Preview card
        st.markdown("<div class='audit-card'>", unsafe_allow_html=True)
        _render_step_header("02", "File Analysis", "Detected questions per sheet")
        preview_rows, total_questions = _build_preview_rows(parsed)
        _render_preview_table(preview_rows)

        st.markdown(
                f"""
                <div style="margin-top:16px;display:flex;align-items:flex-end;gap:10px;">
                    <div style="font-family:'DM Mono',monospace;font-size:34px;color:#F1F0FF;line-height:1;">{total_questions}</div>
                    <div style="color:#9898B0;font-size:13px;padding-bottom:6px;">questions detected</div>
                </div>
                """,
                unsafe_allow_html=True,
        )

        st.markdown("</div>", unsafe_allow_html=True)

    if total_questions == 0:
        _render_alert("warning", "No questions detected", "Try another file or confirm the workbook contains an audit questionnaire.")
        return

    if not (api_key or "").strip():
        st.markdown("<div class='audit-card'>", unsafe_allow_html=True)
        _render_step_header("03", "Generate AI Answers", "Add your key to start processing")
        _render_alert("warning", "API key required", "Enter your Anthropic API key in the sidebar to generate answers.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    # STEP 3 — Process card
    st.markdown("<div class='audit-card'>", unsafe_allow_html=True)
    _render_step_header("03", "Generate AI Answers", "Claude answers each question using standard audit language")
    generate = st.button("Generate Answers", type="primary")

    if not generate:
        st.markdown("</div>", unsafe_allow_html=True)
        return

    progress = st.progress(0)
    status = st.empty()
    current = st.empty()
    eta_box = st.empty()
    spinner = st.empty()

    answered_items: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []

    for idx, q in enumerate(parsed, start=1):
        sheet = str(q.get("sheet") or "")
        question_text = str(q.get("question") or "").strip()
        criteria = q.get("criteria") if isinstance(q.get("criteria"), dict) else {}

        status.text(f"Processing question {idx} of {total_questions}...")
        current.write(
            {
                "Current sheet": sheet,
                "Current question": (question_text[:250] + "...") if len(question_text) > 250 else question_text,
            }
        )

        eta_seconds = _estimate_seconds_remaining(total_questions, idx - 1)
        eta_box.markdown(
            f"<div style='color:#9898B0;font-size:12px;'>Estimated time remaining: <span style='font-family:DM Mono,monospace;color:#F1F0FF;'>{_format_seconds(eta_seconds)}</span></div>",
            unsafe_allow_html=True,
        )
        spinner.markdown("<div class='spin'>⟳</div>", unsafe_allow_html=True)

        try:
            ai = answer_question(question_text, criteria)
        except Exception as exc:
            skipped.append(
                {
                    "sheet": sheet,
                    "row_index": q.get("row_index"),
                    "sl": q.get("sl"),
                    "question": question_text,
                    "reason": str(exc),
                }
            )
            progress.progress(int(idx / total_questions * 100))
            time.sleep(0.5)
            continue

        # If AI returned empty data, treat as skipped
        score = str(ai.get("score") or "").strip()
        present_status = str(ai.get("present_status") or "").strip()
        improvement_plan = str(ai.get("improvement_plan") or "").strip()

        if not (score or present_status or improvement_plan):
            skipped.append(
                {
                    "sheet": sheet,
                    "row_index": q.get("row_index"),
                    "sl": q.get("sl"),
                    "question": question_text,
                    "reason": "Empty AI answer",
                }
            )
            progress.progress(int(idx / total_questions * 100))
            time.sleep(0.5)
            continue

        answered_items.append(
            {
                **q,
                "score": score,
                "present_status": present_status,
                "improvement_plan": improvement_plan,
            }
        )

        progress.progress(int(idx / total_questions * 100))
        time.sleep(0.5)

    spinner.empty()
    st.markdown("</div>", unsafe_allow_html=True)

    # Write workbook
    write_error: Optional[Exception] = None
    write_result: Optional[WriteResult] = None

    try:
        write_result = write_answers(str(input_path), answered_items)
        st.session_state["output_path"] = write_result.output_path
    except Exception as exc:
        write_error = exc

    # Update sidebar session stats
    st.session_state["files_processed"] = int(st.session_state.get("files_processed", 0) or 0) + 1
    st.session_state["questions_answered"] = int(st.session_state.get("questions_answered", 0) or 0) + len(answered_items)

    # STEP 4 — Results card
    st.markdown("<div class='audit-card'>", unsafe_allow_html=True)
    _render_step_header("04", "Results", "Summary of this processing run")

    answered_count = len(answered_items)
    skipped_count = len(skipped)

    score_counts = Counter(_score_bucket(item.get("score", "")) for item in answered_items)

    st.markdown(
        f"""
        <div style="background:#0D2518;border:1px solid #10B981;border-radius:12px;padding:12px 14px;margin-bottom:16px;">
          <div style="color:#10B981;font-size:13px;font-weight:500;">✓ Processing complete — {answered_count} questions answered</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(
            f"""
            <div style="background:#16161E;border:1px solid #2A2A3A;border-radius:12px;padding:16px;">
              <div style="font-family:'DM Mono',monospace;font-size:22px;color:#F1F0FF;">{total_questions}</div>
              <div style="color:#9898B0;font-size:12px;">Total</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            f"""
            <div style="background:#16161E;border:1px solid #2A2A3A;border-radius:12px;padding:16px;">
              <div style="font-family:'DM Mono',monospace;font-size:22px;color:#10B981;">{answered_count}</div>
              <div style="color:#9898B0;font-size:12px;">Answered</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            f"""
            <div style="background:#16161E;border:1px solid #2A2A3A;border-radius:12px;padding:16px;">
              <div style="font-family:'DM Mono',monospace;font-size:22px;color:#6366F1;">{int(score_counts.get('10', 0))}</div>
              <div style="color:#9898B0;font-size:12px;">Score 10</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c4:
        st.markdown(
            f"""
            <div style="background:#16161E;border:1px solid #2A2A3A;border-radius:12px;padding:16px;">
              <div style="font-family:'DM Mono',monospace;font-size:22px;color:#F59E0B;">{skipped_count}</div>
              <div style="color:#9898B0;font-size:12px;">Skipped</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<div style='height:14px;'></div>", unsafe_allow_html=True)

    # Score breakdown bar chart
    chart_data = {
        "10": int(score_counts.get("10", 0)),
        "5": int(score_counts.get("5", 0)),
        "3": int(score_counts.get("3", 0)),
        "0": int(score_counts.get("0", 0)),
    }
    st.bar_chart(chart_data)

    if skipped:
        _render_alert("warning", "Some questions were skipped", "Review the list below (API/format issues).")
        st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
        st.dataframe(skipped, use_container_width=True)

    if write_error is not None or write_result is None:
        _render_alert("error", "Write failed", f"Failed to write filled Excel output.\n{write_error}")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    st.markdown(
        """
        <div style="margin-top:14px;margin-bottom:10px;color:#9898B0;font-size:12px;">
          Output is ready for download.
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Download button
    try:
        out_path = Path(write_result.output_path)
        data = out_path.read_bytes()
        st.markdown("<div class='download-button'>", unsafe_allow_html=True)
        st.download_button(
            label="⬇️ Download Filled Excel",
            data=data,
            file_name=out_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.markdown("</div>", unsafe_allow_html=True)
    except Exception as exc:
        _render_alert("error", "Download failed", f"Output file created, but download failed.\n{exc}")

    st.markdown("</div>", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
